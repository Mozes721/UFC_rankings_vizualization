import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
import numpy as np
import xlrd
import re

#get all sheet names
sheet_names = []
wb = load_workbook(filename = 'UFC_rankings.xlsx')

for sheet in wb.sheetnames:
    sheet_names.append(sheet)

print(sheet_names)

mpp = pd.read_excel("UFC_rankings.xlsx", sheet_name="Men's pound-for-pound")

#get just the fighter and Record
menspp = mpp.iloc[:, 1:3]
#menspp.Record = menspp.Record.str.replace('-', ' ')

menspp_10 = menspp.head(11)
#change to Dataframe
mendf = pd.DataFrame(menspp_10)

#adjust record values in dataframe
mendf = mendf.applymap(lambda x: re.sub("\(.*?\)", "", x))
mendf.head()

#bar/pie plot
figher_array = []
record = []
#assign dataframe values in its corresponding array variable
for index, row in mendf.iterrows():
    #print(row["Fighter"], row["Record"])
    figher_array.append(row["Fighter"])
    record.append(row["Record"])

#bar/pie plot
figher_array = []
record = []
#assign dataframe values in its corresponding array variable
for index, row in mendf.iterrows():
    #print(row["Fighter"], row["Record"])
    figher_array.append(row["Fighter"])
    record.append(row["Record"])

records = [x.strip(' ') for x in record]
wins = []
losses = []
for won in records:
    wins.append(won[0:2])
for loss in records:
    losses.append(loss[-1])


#change to integers
record_wins = [int(x) for x in wins]
record_losses = [int(x) for x in losses]

#no iterateors are passed
record = zip()

#Converting iterator to list
result_record = list(record)

#two iterables are passed in
record = zip(record_wins, record_losses)

# Converting iterator to set
fighter_records = list(record)

ypos = np.arange(len(figher_array))
xpos = [0, 1, 2, 3, 4, 5, 6, 7, 8]
plt.rcParams.update({'font.size': 22})

#bar plot
f, ax = plt.subplots(figsize=(22,5))
plt.title(sheet_names[0])
plt.xticks(rotation='vertical')
plt.xlabel("Fighters")
plt.ylabel("UFC Record")
plt.xticks(ypos, figher_array)
plt.bar(ypos - 0.2, record_wins, 0.4, label = 'Wins', color='g')
plt.bar(ypos + 0.2, record_losses, 0.4, label = 'Losses', color='r')
plt.legend()
plt.show()