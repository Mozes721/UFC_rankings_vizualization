from openpyxl import load_workbook
from bar_chart import bar_graph
from pie_chart import pie_chart

#get all sheet names
sheet_names = []
wb = load_workbook(filename = 'UFC_rankings.xlsx')

for sheet in wb.sheetnames:
    sheet_names.append(sheet)

print(sheet_names)

def choose_sheet():
    sheet_input = input("Please choose one of the fighter rankings for displaying bar chart: ")
    running = True
    while running:
        if sheet_input in sheet_names:
            print("Sheetname found:")
            print("1: Check it as a bar graph?")
            print("2: Check for fighters individual stats as a pie chart?")
            print(" ")
            bar_or_pie  = int(input("Input: "))
            if bar_or_pie == 1:
                bar_graph(sheet_input)
            elif bar_or_pie == 2:
                pie_chart(sheet_input)
            else:
                print("wrong selection...")
                break
        else:
            print("Try again...")
            choose_sheet()
        running = False

if __name__ == '__main__':
    choose_sheet()